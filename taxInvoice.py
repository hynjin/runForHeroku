#!/usr/bin/env python
# coding: utf-8
import openpyxl
import os
import datetime
from io import StringIO,BytesIO

#erp, email : 엑셀파일
def tax(erp,email):
    erpWB = openpyxl.load_workbook(erp)
    emailWB =  openpyxl.load_workbook(email)
    taxWB = openpyxl.load_workbook('data/origin.xlsx')

    erpWS = erpWB[erpWB.sheetnames[0]]
    emailWS = emailWB.active
    taxWS = taxWB[taxWB.sheetnames[0]]

    # 이메일 사업자 번호를 키로 해쉬테이블
    emailDic = {'사업자번호':0, '이메일주소':1}
    emailHash = dict()
    for i in range (0,emailWS.max_column):
        emailHash[emailWS[1][i].value] = emailWS[1][i].coordinate.strip('0123456789') #coordinate:셀 좌표
        print(emailWS[1][i].value,emailHash[emailWS[1][i].value])
    emailAddress = dict()
    for i in range(2,emailWS.max_row+1):
        if emailWS[emailHash['사업자번호']+str(i)].value!=None:
            etemp = emailWS[emailHash['사업자번호']+str(i)].value.strip(' ')
        if(etemp != ''):
            emailAddress[int(etemp)] = emailWS[emailHash['이메일주소']+str(i)].value
        print(emailWS[emailHash['이메일주소']+str(i)].value,emailWS[emailHash['사업자번호']+str(i)].value)

    # erp 수금 기록 열 이름
    # 사업자번호, 법인명, 대표자, 업태, 종목, 사업장주소, 공급가액, 세액, 수금일자 사용
    erpDic = {'사업자번호':0,'법인명':1,'대표자':2,'업태':3,'종목':4,'사업장 주소':5,'공급가액':6,'세액':7, '수금일자':8}
    erpHash = dict() #열 이름을 키값으로 갖는 셀 위치
    for c in range (0,erpWS.max_column):
        erpHash[erpWS[2][c].value] = erpWS[2][c].coordinate.strip('0123456789')
        #print(erpWS[2][c].value,erpHash[erpWS[2][c].value])

    erpVal = []
    for i in range (3,erpWS.max_row+1):
        temp = []
        if erpWS[erpHash['사업자번호']+str(i)].value!=None and erpWS[erpHash['수금액']+str(i)].value!=None: #사업자명이나 수금액이 없다면 저장하지 않음
            for j in erpDic:
                if j=='대표자': #대표자 이름 전처리
                    temp.append(erpWS[erpHash[j]+str(i)].value.strip(' '))
                else:
                    temp.append(erpWS[erpHash[j]+str(i)].value)
            if temp[0] in emailAddress.keys(): #저장된 이메일 주소가 없을 경우 None
                temp.append(emailAddress[temp[0]])
            else:
                temp.append(None)
            erpVal.append(temp)
            #print(temp)

    # 전자세금계산서 저장을 위한 이차원 배열
    rowVal = []
    dt = datetime.datetime.now()
    for i in range (len(erpVal)):
        temp = []
        for j in range (taxWS.max_column):
            if j==0:
                temp.append('01')
            elif j==1:
                temp.append(erpVal[i][erpDic['수금일자']])
            elif j==2:
                temp.append(erpVal[i][erpDic['사업자번호']])
            elif j==4:
                temp.append(erpVal[i][erpDic['법인명']])
            elif j==5:
                temp.append(erpVal[i][erpDic['대표자']])
            elif j==6:
                temp.append(erpVal[i][erpDic['사업장 주소']])
            elif j==7:
                temp.append(erpVal[i][erpDic['업태']])
            elif j==8:
                temp.append(erpVal[i][erpDic['종목']])
            elif j==9:
                temp.append(erpVal[i][9]) #이메일
            elif j==11:
                temp.append(erpVal[i][erpDic['공급가액']])
            elif j==12:
                temp.append(erpVal[i][erpDic['세액']])
            elif j==14:
                num = erpVal[i][erpDic['수금일자']]%100
                if num < 10:
                    num = '0' + str(num)
                temp.append(num)
            elif j==15:
                month = (int(dt.strftime("%m")) + 11)%12
                temp.append(str(month) +'월 CCTV용역료')
            elif j==19:
                temp.append(erpVal[i][erpDic['공급가액']])
            elif j==20:
                temp.append(erpVal[i][erpDic['세액']])
            elif j==50:
                temp.append('01') #영수01 청구02
            else:
                temp.append('')
        rowVal.append(temp)
        #print(temp)

    for i in range (len(rowVal)):
        for j in range (taxWS.max_column):
            taxWS.cell(i+7,j+1,rowVal[i][j]).font = openpyxl.styles.Font(color="000000")
            #print(rowVal[i][j])

    # taxWB.save('data/test.xlsx')

    virtual_workbook = BytesIO()
    taxWB.save(virtual_workbook)

    erpWB.close()
    emailWB.close()
    taxWB.close()

    return virtual_workbook
